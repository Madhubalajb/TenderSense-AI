"""
utils/excel_export.py — Excel Report Generator
-------------------------------------------------
Generates a multi-sheet Excel evaluation report using openpyxl.

Sheets:
  1. Summary       — one row per bidder, overall verdicts
  2. Criterion Detail — one row per (bidder × criterion) pair
  3. Criteria List — the locked criteria from the tender
  4. Audit Log     — all recorded events

Designed to be:
  - Sortable and filterable by procurement staff
  - Importable into government MIS systems
  - Readable without any special software (standard .xlsx)

Library: openpyxl (open-source, pure Python)
Install: pip install openpyxl
"""

import io
import os
import json
import time
from typing import Optional

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as XlTable, TableStyleInfo


# ── Colour constants ───────────────────────────────────────────────────────────
NAVY     = "1F4E79"
BLUE     = "2E75B6"
LIGHT    = "DEEAF1"
GREEN    = "375623"
GREEN_BG = "E2EFDA"
RED      = "C00000"
RED_BG   = "FCE4D6"
AMBER    = "7F6000"
AMBER_BG = "FFF2CC"
GREY_BG  = "F5F5F5"
WHITE    = "FFFFFF"

# ── Border styles ──────────────────────────────────────────────────────────────
_thin   = Side(style="thin",   color="CCCCCC")
_medium = Side(style="medium", color="2E75B6")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_header_border = Border(
    left=_medium, right=_medium, top=_medium, bottom=_medium
)


# ── Style factories ────────────────────────────────────────────────────────────
def _header_style(bg_color: str = BLUE) -> dict:
    return {
        "font":   Font(bold=True, color=WHITE, size=9),
        "fill":   PatternFill("solid", fgColor=bg_color),
        "align":  Alignment(horizontal="center", vertical="center",
                             wrap_text=True),
        "border": _border,
    }

def _cell_style(bg: str = WHITE, color: str = "1A1A1A",
                bold: bool = False, wrap: bool = True) -> dict:
    return {
        "font":   Font(color=color, size=9, bold=bold),
        "fill":   PatternFill("solid", fgColor=bg),
        "align":  Alignment(vertical="top", wrap_text=wrap),
        "border": _border,
    }

def _apply(cell, style: dict):
    """Apply a style dict to an openpyxl cell."""
    if "font"   in style: cell.font      = style["font"]
    if "fill"   in style: cell.fill      = style["fill"]
    if "align"  in style: cell.alignment = style["align"]
    if "border" in style: cell.border    = style["border"]


def _verdict_colours(verdict: str) -> tuple:
    """Return (bg_hex, text_hex) for a verdict string."""
    return {
        "pass":       (GREEN_BG, GREEN),
        "eligible":   (GREEN_BG, GREEN),
        "fail":       (RED_BG,   RED),
        "ineligible": (RED_BG,   RED),
        "review":     (AMBER_BG, AMBER),
    }.get(verdict.lower(), (WHITE, "1A1A1A"))


# ── Main generate function ─────────────────────────────────────────────────────

def generate_excel(
    all_verdicts:  list,
    criteria:      list,
    meta:          dict,
    audit_entries: Optional[list] = None,
) -> bytes:
    """
    Generate a complete Excel evaluation report.

    Args:
        all_verdicts:  List of bidder verdict dicts from verdict.py
        criteria:      List of criterion dicts from criteria_llm.py
        meta: {
            "tender_ref":   "CRPF/HQ/2024/087",
            "tender_name":  "tender.pdf",
            "eval_date":    "01 January 2025",
            "department":   "CRPF Procurement Directorate",
            "prepared_by":  "Priya Sharma",
        }
        audit_entries: Optional list of audit log dicts

    Returns:
        Excel file as bytes
    """
    wb = openpyxl.Workbook()

    # Build all sheets
    _build_summary_sheet(wb, all_verdicts, meta)
    _build_detail_sheet(wb, all_verdicts)
    _build_criteria_sheet(wb, criteria)
    _build_audit_sheet(wb, audit_entries or _load_audit_entries())

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── Sheet 1: Summary ───────────────────────────────────────────────────────────

def _build_summary_sheet(wb, all_verdicts: list, meta: dict):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    # ── Title block ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value     = "Tendra AI — Tender Eligibility Evaluation Report"
    title_cell.font      = Font(bold=True, size=14, color=WHITE)
    title_cell.fill      = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Metadata rows
    meta_rows = [
        ("Tender Reference:", meta.get("tender_ref", "—")),
        ("Document:",         meta.get("tender_name", "—")),
        ("Department:",       meta.get("department", "—")),
        ("Evaluation Date:",  meta.get("eval_date", "—")),
        ("Prepared by:",      meta.get("prepared_by", "—")),
        ("Generated at:",     time.strftime("%d %B %Y %H:%M")),
    ]
    for i, (label, value) in enumerate(meta_rows, start=2):
        ws.merge_cells(f"A{i}:B{i}")
        ws.merge_cells(f"C{i}:G{i}")
        label_cell = ws.cell(row=i, column=1, value=label)
        value_cell = ws.cell(row=i, column=3, value=value)
        label_cell.font = Font(bold=True, size=9, color=NAVY)
        value_cell.font = Font(size=9)
        label_cell.fill = PatternFill("solid", fgColor=LIGHT)
        value_cell.fill = PatternFill("solid", fgColor=GREY_BG)

    # ── Metrics row ────────────────────────────────────────────────────────────
    total      = len(all_verdicts)
    eligible   = sum(1 for v in all_verdicts if v["overall_verdict"] == "eligible")
    ineligible = sum(1 for v in all_verdicts if v["overall_verdict"] == "ineligible")
    review     = sum(1 for v in all_verdicts if v["overall_verdict"] == "review")
    avg_conf   = (
        sum(v["overall_confidence"] for v in all_verdicts) / total
        if total else 0
    )

    metric_row = 9
    ws.row_dimensions[metric_row].height = 36

    metrics = [
        ("Total", total,            BLUE),
        ("Eligible", eligible,      GREEN_BG),
        ("Ineligible", ineligible,  RED_BG),
        ("Needs Review", review,    AMBER_BG),
        ("Avg Confidence", f"{avg_conf:.0%}", LIGHT),
    ]
    for col, (label, value, bg) in enumerate(metrics, start=1):
        ws.merge_cells(
            start_row=metric_row, start_column=col * 2 - 1,
            end_row=metric_row,   end_column=col * 2,
        )
        # Label cell (above)
        label_cell = ws.cell(row=metric_row - 1, column=col * 2 - 1, value=label)
        label_cell.font      = Font(bold=True, size=8, color=NAVY)
        label_cell.alignment = Alignment(horizontal="center")
        # Value cell
        val_cell = ws.cell(row=metric_row, column=col * 2 - 1, value=value)
        val_cell.font      = Font(bold=True, size=18, color=NAVY)
        val_cell.fill      = PatternFill("solid", fgColor=bg)
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        val_cell.border    = _border

    # ── Summary table ──────────────────────────────────────────────────────────
    header_row = 12
    headers = ["#", "Bidder Name", "Overall Verdict",
               "Passed", "Failed", "Review", "Confidence"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        _apply(cell, _header_style())
    ws.row_dimensions[header_row].height = 20

    for row_i, v in enumerate(all_verdicts, start=header_row + 1):
        ov          = v["overall_verdict"]
        bg, txt     = _verdict_colours(ov)
        row_bg      = GREY_BG if (row_i - header_row) % 2 == 0 else WHITE

        row_data = [
            row_i - header_row,
            v["bidder_name"],
            ov.upper(),
            v["passed"],
            v["failed"],
            v["review_needed"],
            f"{v['overall_confidence']:.0%}",
        ]
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_i, column=col, value=value)
            style = _cell_style(bg=row_bg if col != 3 else bg,
                                 color="1A1A1A" if col != 3 else txt)
            if col == 3:
                style["font"] = Font(bold=True, size=9, color=txt)
            _apply(cell, style)

        ws.row_dimensions[row_i].height = 18

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 16
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 13


# ── Sheet 2: Criterion Detail ──────────────────────────────────────────────────

def _build_detail_sheet(wb, all_verdicts: list):
    ws = wb.create_sheet("Criterion Detail")
    ws.sheet_view.showGridLines = False

    headers = [
        "Bidder", "Criterion ID", "Criterion Text",
        "Category", "Mandatory", "Verdict", "Confidence",
        "Evidence", "Reasoning", "Method", "Officer Override"
    ]
    col_widths = [25, 10, 42, 12, 10, 12, 12, 40, 40, 12, 35]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        _apply(cell, _header_style())
        ws.column_dimensions[get_column_letter(col)].width = col_widths[col - 1]

    row_i = 2
    for v in all_verdicts:
        for r in v.get("criteria_results", []):
            override = r.get("officer_override", {})
            override_str = (
                f"By {override.get('officer', '')} "
                f"at {override.get('timestamp', '')}: "
                f"{override.get('comment', '')}"
                if override else ""
            )
            verdict = r.get("verdict", "review")
            bg, txt = _verdict_colours(verdict)
            row_bg  = GREY_BG if row_i % 2 == 0 else WHITE

            row_data = [
                v["bidder_name"],
                r.get("criterion_id", ""),
                r.get("criterion_text", ""),
                r.get("category", ""),
                "Yes" if r.get("mandatory") else "No",
                verdict.upper(),
                f"{r.get('confidence', 0):.0%}",
                r.get("evidence", ""),
                r.get("reasoning", ""),
                r.get("layer_used", "").upper(),
                override_str,
            ]
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_i, column=col, value=value)
                if col == 6:
                    _apply(cell, _cell_style(bg=bg, color=txt, bold=True))
                else:
                    _apply(cell, _cell_style(bg=row_bg))
            ws.row_dimensions[row_i].height = 30
            row_i += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


# ── Sheet 3: Criteria List ─────────────────────────────────────────────────────

def _build_criteria_sheet(wb, criteria: list):
    ws = wb.create_sheet("Criteria List")
    ws.sheet_view.showGridLines = False

    headers = ["ID", "Criterion Text", "Category",
               "Mandatory", "Threshold", "Years", "Section"]
    col_widths = [7, 55, 14, 11, 16, 8, 20]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        _apply(cell, _header_style())
        ws.column_dimensions[get_column_letter(col)].width = col_widths[col - 1]

    for row_i, c in enumerate(criteria, start=2):
        row_bg = GREY_BG if row_i % 2 == 0 else WHITE
        row_data = [
            c.get("id", ""),
            c.get("criterion_text", ""),
            c.get("category", ""),
            "MANDATORY" if c.get("mandatory") else "Optional",
            c.get("threshold_value", ""),
            c.get("years_required", ""),
            c.get("section_reference", ""),
        ]
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_i, column=col, value=value)
            style = _cell_style(bg=row_bg)
            if col == 4 and c.get("mandatory"):
                style = _cell_style(bg=RED_BG, color=RED, bold=True)
            _apply(cell, style)
        ws.row_dimensions[row_i].height = 25

    ws.freeze_panes = "A2"


# ── Sheet 4: Audit Log ─────────────────────────────────────────────────────────

def _build_audit_sheet(wb, audit_entries: list):
    ws = wb.create_sheet("Audit Log")
    ws.sheet_view.showGridLines = False

    headers = ["Timestamp", "User / Officer", "Event", "Detail"]
    col_widths = [20, 25, 22, 60]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        _apply(cell, _header_style(bg_color=NAVY))
        ws.column_dimensions[get_column_letter(col)].width = col_widths[col - 1]

    for row_i, entry in enumerate(audit_entries, start=2):
        row_bg = GREY_BG if row_i % 2 == 0 else WHITE
        row_data = [
            str(entry.get("timestamp", ""))[:19].replace("T", " "),
            entry.get("user", ""),
            entry.get("event", ""),
            entry.get("detail", ""),
        ]
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_i, column=col, value=str(value))
            _apply(cell, _cell_style(bg=row_bg))
        ws.row_dimensions[row_i].height = 16

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


# ── Load audit entries from JSONL file ─────────────────────────────────────────

def _load_audit_entries() -> list:
    """Load audit log entries from the JSONL file."""
    path = "data/uploads/audit_log.jsonl"
    if not os.path.exists(path):
        return []
    entries = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    entries.append(json.loads(line))
                except json.JSONDecodeError:
                    pass
    return list(reversed(entries))


# ── Quick test ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    sample = [{
        "bidder_name":        "Test Bidder Pvt Ltd",
        "overall_verdict":    "eligible",
        "overall_confidence": 0.91,
        "passed": 2, "failed": 0, "review_needed": 0,
        "total_criteria": 2,
        "summary": "All criteria met.",
        "criteria_results": [
            {
                "criterion_id":   "C001",
                "criterion_text": "Annual turnover ≥ ₹2 Crore",
                "category":       "Financial",
                "mandatory":      True,
                "verdict":        "pass",
                "confidence":     0.92,
                "evidence":       "₹3.4 Cr found",
                "reasoning":      "Exceeds threshold",
                "layer_used":     "numeric",
            },
        ],
    }]

    xlsx_bytes = generate_excel(
        all_verdicts=sample,
        criteria=[{
            "id": "C001", "criterion_text": "Annual turnover ≥ ₹2 Crore",
            "category": "Financial", "mandatory": True,
        }],
        meta={
            "tender_ref":  "CRPF/TEST/2024/001",
            "tender_name": "test.pdf",
            "eval_date":   "01 January 2025",
            "department":  "CRPF",
            "prepared_by": "Test Officer",
        }
    )

    with open("test_report.xlsx", "wb") as f:
        f.write(xlsx_bytes)
    print(f"Test Excel generated: test_report.xlsx ({len(xlsx_bytes):,} bytes)")