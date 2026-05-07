"""
export.py — Report Generation Page
-------------------------------------
Generates the final evaluation report in two formats:
  1. PDF  — formatted, printable, suitable for official submission
  2. Excel — machine-readable, suitable for record-keeping and further analysis

Both reports include:
  - Tender reference and date
  - Summary table (all bidders, overall verdict)
  - Criterion-level detail for every bidder
  - Officer overrides with names and timestamps
  - Full audit trail

Libraries used:
  - ReportLab: PDF generation (open-source, pure Python)
  - openpyxl:  Excel generation (open-source, pure Python)
"""

import io
import os
import time

import streamlit as st

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RL_GREEN    = colors.HexColor("#375623")
RL_RED      = colors.HexColor("#C00000")
RL_AMBER    = colors.HexColor("#7F6000")
RL_NAVY     = colors.HexColor("#0F172A")
RL_TEAL     = colors.HexColor("#0D9488")
RL_HEADER   = colors.HexColor("#1E293B")
RL_LIGHT    = colors.HexColor("#F0FDFA")
RL_GREEN_BG = colors.HexColor("#ECFDF5")
RL_RED_BG   = colors.HexColor("#FEF2F2")
RL_AMBER_BG = colors.HexColor("#FFFBEB")


def show():
    st.title("Export")
    st.caption("Generate the final evaluation report in PDF and Excel formats.")

    all_verdicts = st.session_state.get("all_verdicts", [])
    if not all_verdicts:
        st.warning("No evaluation results to export. Complete the evaluation first.")
        st.stop()

    tender_name  = st.session_state.get("tender_filename", "Tender")
    officer      = st.session_state.get("officer_name", "Officer")
    criteria     = st.session_state.get("criteria", [])

    review_count = sum(
        sum(1 for r in v["criteria_results"]
            if r["verdict"] == "review" and "officer_override" not in r)
        for v in all_verdicts
    )

    if review_count > 0:
        st.warning(
            f"There are **{review_count} criterion/criteria** still flagged for officer review. "
            "You can export now, but these will appear as 'Pending Review' in the report."
        )

    st.divider()
    st.subheader("Report Details")

    col1, col2 = st.columns(2)
    with col1:
        tender_ref   = st.text_input("Tender reference number",
                                      value="CRPF/HQ/2024/001",
                                      placeholder="e.g. CRPF/HQ/2024/087")
        eval_date    = st.text_input("Evaluation date", value=time.strftime("%d %B %Y"))
    with col2:
        dept         = st.text_input("Department / Organisation",
                                      value="CRPF Procurement Directorate")
        prepared_by  = st.text_input("Report prepared by", value=officer)

    report_meta = {
        "tender_ref":  tender_ref,
        "tender_name": tender_name,
        "eval_date":   eval_date,
        "department":  dept,
        "prepared_by": prepared_by,
    }

    st.divider()

    locked = st.checkbox(
        f"I confirm that I have reviewed all evaluation results and this report is ready for "
        f"submission. By checking this box, I am signing off as **{officer}** on "
        f"{time.strftime('%d %B %Y at %H:%M')}.",
        key="report_locked",
    )

    st.divider()

    col_pdf, col_excel = st.columns(2)

    with col_pdf:
        st.markdown(
            '<div style="background:#F8FAFC;border:1px solid #E2E8F0;border-radius:8px;'
            'padding:16px 18px;margin-bottom:12px;">'
            '<div style="font-size:13px;font-weight:600;color:#1E293B;margin-bottom:4px;">'
            'PDF Report</div>'
            '<div style="font-size:12px;color:#64748B;">Formatted, printable, suitable for '
            'official submission and record-keeping.</div></div>',
            unsafe_allow_html=True,
        )
        if st.button("Generate PDF Report", type="primary",
                     use_container_width=True, disabled=not locked):
            with st.spinner("Generating PDF..."):
                pdf_bytes = generate_pdf(all_verdicts, criteria, report_meta)
            st.download_button(
                label="Download PDF",
                data=pdf_bytes,
                file_name=f"Tendra_Evaluation_{tender_ref.replace('/', '-')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
            _log_export("pdf", tender_ref, officer)

    with col_excel:
        st.markdown(
            '<div style="background:#F8FAFC;border:1px solid #E2E8F0;border-radius:8px;'
            'padding:16px 18px;margin-bottom:12px;">'
            '<div style="font-size:13px;font-weight:600;color:#1E293B;margin-bottom:4px;">'
            'Excel Report</div>'
            '<div style="font-size:12px;color:#64748B;">Machine-readable, suitable for '
            'further analysis and audit trail review.</div></div>',
            unsafe_allow_html=True,
        )
        if st.button("Generate Excel Report", use_container_width=True, disabled=not locked):
            with st.spinner("Generating Excel..."):
                excel_bytes = generate_excel(all_verdicts, criteria, report_meta)
            st.download_button(
                label="Download Excel",
                data=excel_bytes,
                file_name=f"Tendra_Evaluation_{tender_ref.replace('/', '-')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            _log_export("excel", tender_ref, officer)

    if not locked:
        st.caption("Check the confirmation box above to enable report generation.")


def generate_pdf(all_verdicts, criteria, meta):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )
    styles = getSampleStyleSheet()
    story  = []

    title_style = ParagraphStyle(
        "CustomTitle", parent=styles["Title"],
        textColor=RL_NAVY, fontSize=18, spaceAfter=6,
    )
    heading_style = ParagraphStyle(
        "CustomHeading", parent=styles["Heading2"],
        textColor=RL_TEAL, fontSize=12, spaceBefore=12, spaceAfter=4,
    )
    small_style = ParagraphStyle("Small", parent=styles["Normal"], fontSize=8)

    story.append(Paragraph("Tendra AI", title_style))
    story.append(Paragraph("Tender Eligibility Evaluation Report", styles["Heading2"]))
    story.append(HRFlowable(width="100%", thickness=1, color=RL_NAVY, spaceAfter=10))

    meta_data = [
        ["Tender Reference:", meta.get("tender_ref", "")],
        ["Tender File:",      meta.get("tender_name", "")],
        ["Department:",       meta.get("department", "")],
        ["Evaluation Date:",  meta.get("eval_date", "")],
        ["Prepared by:",      meta.get("prepared_by", "")],
        ["Generated at:",     time.strftime("%d %B %Y %H:%M:%S")],
    ]
    meta_table = Table(meta_data, colWidths=[5*cm, 11*cm])
    meta_table.setStyle(TableStyle([
        ("FONTSIZE",    (0, 0), (-1, -1), 9),
        ("FONTNAME",    (0, 0), (0, -1), "Helvetica-Bold"),
        ("TEXTCOLOR",   (0, 0), (0, -1), RL_NAVY),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("TOPPADDING",  (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 0.5*cm))

    story.append(Paragraph("Evaluation Summary", heading_style))
    summary_header = [["Bidder", "Overall Verdict", "Passed", "Failed", "Review", "Confidence"]]
    summary_rows   = []
    for v in all_verdicts:
        summary_rows.append([
            v["bidder_name"], v["overall_verdict"].upper(),
            str(v["passed"]), str(v["failed"]),
            str(v["review_needed"]), f"{v['overall_confidence']:.0%}",
        ])

    summary_table = Table(
        summary_header + summary_rows,
        colWidths=[6*cm, 3*cm, 1.5*cm, 1.5*cm, 1.5*cm, 2.5*cm]
    )
    row_styles = [
        ("BACKGROUND",  (0, 0), (-1, 0), RL_HEADER),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for i, v in enumerate(all_verdicts, start=1):
        ov = v["overall_verdict"]
        if ov == "eligible":
            row_styles.append(("BACKGROUND", (1, i), (1, i), RL_GREEN_BG))
            row_styles.append(("TEXTCOLOR",  (1, i), (1, i), RL_GREEN))
        elif ov == "ineligible":
            row_styles.append(("BACKGROUND", (1, i), (1, i), RL_RED_BG))
            row_styles.append(("TEXTCOLOR",  (1, i), (1, i), RL_RED))
        else:
            row_styles.append(("BACKGROUND", (1, i), (1, i), RL_AMBER_BG))
            row_styles.append(("TEXTCOLOR",  (1, i), (1, i), RL_AMBER))
    summary_table.setStyle(TableStyle(row_styles))
    story.append(summary_table)
    story.append(Spacer(1, 0.5*cm))

    story.append(Paragraph("Criterion-Level Detail", heading_style))
    for v in all_verdicts:
        story.append(Paragraph(
            f"Bidder: {v['bidder_name']} — {v['overall_verdict'].upper()}",
            styles["Heading3"]
        ))
        story.append(Paragraph(v["summary"], small_style))
        story.append(Spacer(1, 0.2*cm))

        crit_header = [["ID", "Criterion", "Category", "Verdict", "Confidence", "Evidence"]]
        crit_rows   = []
        for r in v["criteria_results"]:
            crit_rows.append([
                r.get("criterion_id", ""),
                Paragraph(r.get("criterion_text", "")[:120], small_style),
                r.get("category", ""),
                r.get("verdict", "").upper(),
                f"{r.get('confidence', 0):.0%}",
                Paragraph(r.get("evidence", "")[:100], small_style),
            ])

        crit_table = Table(
            crit_header + crit_rows,
            colWidths=[1*cm, 5.5*cm, 2*cm, 1.5*cm, 1.5*cm, 4.5*cm]
        )
        crit_styles = [
            ("BACKGROUND",  (0, 0), (-1, 0), RL_TEAL),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 7.5),
            ("GRID",        (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("TOPPADDING",  (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("VALIGN",      (0, 0), (-1, -1), "TOP"),
        ]
        for i, r in enumerate(v["criteria_results"], start=1):
            verdict = r.get("verdict", "")
            if verdict == "pass":
                crit_styles.append(("TEXTCOLOR", (3, i), (3, i), RL_GREEN))
            elif verdict == "fail":
                crit_styles.append(("TEXTCOLOR", (3, i), (3, i), RL_RED))
            else:
                crit_styles.append(("TEXTCOLOR", (3, i), (3, i), RL_AMBER))
        crit_table.setStyle(TableStyle(crit_styles))
        story.append(crit_table)
        story.append(Spacer(1, 0.4*cm))

    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey, spaceBefore=10))
    story.append(Paragraph(
        "This report was generated by Tendra AI. All verdicts marked as 'REVIEW' require "
        "officer confirmation before being treated as final. Officer overrides are recorded "
        "with name and timestamp for audit purposes.",
        small_style
    ))
    doc.build(story)
    return buffer.getvalue()


def generate_excel(all_verdicts, criteria, meta):
    wb = openpyxl.Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = "Tendra AI — Evaluation Report"
    ws_summary["A1"].font = Font(bold=True, size=14, color="0F172A")
    ws_summary["A2"] = f"Tender: {meta.get('tender_ref', '')} | {meta.get('tender_name', '')}"
    ws_summary["A3"] = f"Date: {meta.get('eval_date', '')} | Prepared by: {meta.get('prepared_by', '')}"
    ws_summary["A3"].font = Font(italic=True, size=10, color="808080")

    headers = ["Bidder Name", "Overall Verdict", "Criteria Passed",
               "Criteria Failed", "Needs Review", "Confidence", "Summary"]
    for col, header in enumerate(headers, start=1):
        cell = ws_summary.cell(row=5, column=col, value=header)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="1E293B")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    fill_green = PatternFill("solid", fgColor="ECFDF5")
    fill_red   = PatternFill("solid", fgColor="FEF2F2")
    fill_amber = PatternFill("solid", fgColor="FFFBEB")

    for row_idx, v in enumerate(all_verdicts, start=6):
        ov = v["overall_verdict"]
        row_data = [
            v["bidder_name"], ov.upper(),
            v["passed"], v["failed"], v["review_needed"],
            f"{v['overall_confidence']:.0%}", v["summary"],
        ]
        for col, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 2:
                cell.fill = (
                    fill_green if ov == "eligible" else
                    fill_red   if ov == "ineligible" else fill_amber
                )

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 16
    ws_summary.column_dimensions["G"].width = 50
    for col in ["C", "D", "E", "F"]:
        ws_summary.column_dimensions[col].width = 14

    ws_detail = wb.create_sheet("Criterion Detail")
    detail_headers = [
        "Bidder", "Criterion ID", "Criterion Text", "Category", "Mandatory",
        "Verdict", "Confidence", "Evidence", "Reasoning", "Method", "Officer Override"
    ]
    for col, header in enumerate(detail_headers, start=1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="0D9488")
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for v in all_verdicts:
        for r in v["criteria_results"]:
            override     = r.get("officer_override")
            override_str = (
                f"By {override['officer']} at {override['timestamp']}: {override['comment']}"
                if override else ""
            )
            row_data = [
                v["bidder_name"], r.get("criterion_id", ""),
                r.get("criterion_text", ""), r.get("category", ""),
                "Yes" if r.get("mandatory") else "No",
                r.get("verdict", "").upper(), f"{r.get('confidence', 0):.0%}",
                r.get("evidence", ""), r.get("reasoning", ""),
                r.get("layer_used", "").upper(), override_str,
            ]
            for col, value in enumerate(row_data, start=1):
                cell = ws_detail.cell(row=row_idx, column=col, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col == 6:
                    verdict = r.get("verdict", "")
                    cell.fill = (
                        fill_green if verdict == "pass" else
                        fill_red   if verdict == "fail" else fill_amber
                    )
            row_idx += 1

    for col, width in zip(
        ["A","B","C","D","E","F","G","H","I","J","K"],
        [25, 8,  40, 12, 10, 10, 10, 35, 35, 10, 35]
    ):
        ws_detail.column_dimensions[col].width = width

    ws_audit = wb.create_sheet("Audit Log")
    ws_audit["A1"] = "Audit Log"
    ws_audit["A1"].font = Font(bold=True, size=12, color="0F172A")
    audit_headers = ["Timestamp", "User", "Event", "Detail"]
    for col, h in enumerate(audit_headers, start=1):
        cell = ws_audit.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E293B")

    try:
        import json
        with open("data/uploads/audit_log.jsonl") as f:
            for row_i, line in enumerate(f, start=3):
                entry = json.loads(line)
                ws_audit.cell(row=row_i, column=1, value=entry.get("timestamp", ""))
                ws_audit.cell(row=row_i, column=2, value=entry.get("officer", ""))
                ws_audit.cell(row=row_i, column=3, value=entry.get("event", ""))
                ws_audit.cell(row=row_i, column=4, value=entry.get("comment", ""))
    except FileNotFoundError:
        ws_audit.cell(row=3, column=1, value="No audit events recorded yet")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _log_export(format_type, tender_ref, officer):
    import json
    os.makedirs("data/uploads", exist_ok=True)
    with open("data/uploads/audit_log.jsonl", "a") as f:
        f.write(json.dumps({
            "event": "report_exported", "format": format_type,
            "tender": tender_ref, "officer": officer,
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
        }) + "\n")


if __name__ == "__main__":
    st.set_page_config(page_title="Tendra AI — Export", layout="wide")
    show()